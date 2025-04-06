from flask import Flask, render_template, request, send_from_directory
import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

app = Flask(__name__)


# Ruta para la carga de archivos y procesamiento
@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        file = request.files["txt_file"]
        if file:

            if not os.path.exists("uploads"):
                os.makedirs("uploads")

            file_path = os.path.join("uploads", file.filename)
            file.save(file_path)

            try:
                try:
                    with open(file_path, "r", encoding="utf-8") as f:
                        lines = f.readlines()
                except UnicodeDecodeError:
                    with open(file_path, "r", encoding="latin-1") as f:
                        lines = f.readlines()

                # Limpiar caracteres de control y eliminar bloques entre "----" y "--"
                cleaned_lines = []
                movements = []
                eliminar = False
                eliminar_desde_totales = False
                temp_movement = {}
                compras_o_ventas = ""
                jurisdicciones = set()

                # CONCEPTOS

                conceptos = [
                    {"1": "Mercaderia c/iva"},
                    {"2": "mercaderia s/iva"},
                    {"3": "perecederos"},
                    {"4": "carnes"},
                    {"5": "verduras"},
                    {"6": "huevos"},
                    {"7": "pollos"},
                    {"8": "no perecederos"},
                    {"9": "materia prima c/iva"},
                    {"10": "materia prima s/iva"},
                    {"11": "materiales c/iva"},
                    {"12": "materiales s/iva"},
                    {"13": "productos varios"},
                    {"14": "alimentos balanceados"},
                    {"15": "bienes de cambio"},
                    {"16": "Combustible para la venta"},
                    {"18": "gs de impo/expo"},
                    {"19": "gastos de prestamo"},
                    {"20": "gastos generales c/iva"},
                    {"21": "gastos generales s/iva"},
                    {"22": "gastos bancarios c/iva"},
                    {"23": "gastos bancarios s/iva"},
                    {"24": "gastos adm. C/iva"},
                    {"25": "gastos adm. S/iva"},
                    {"26": "gs.comercializacion c/iva"},
                    {"27": "gs.comercializacion S/iva"},
                    {"28": "servicios varios"},
                    {"29": "imp. Tasas y contribuciones"},
                    {"30": "serv x cta de 3° c/iva"},
                    {"31": "serv x cta de 3° s/iva"},
                    {"32": "gastos despachantes"},
                    {"33": "honorarios c/ivA"},
                    {"34": "honorarios s/iva"},
                    {"35": "derechos de importacion"},
                    {"36": "prestamos"},
                    {"37": "leasing"},
                    {"38": "intereses"},
                    {"39": "gastos de tarjeta"},
                    {"40": "insumos"},
                    {"41": "material de embalaje"},
                    {"42": "seguros comerciales"},
                    {"43": "seguro de vida"},
                    {"44": "seguro de vehiculo"},
                    {"45": "Gastos de vehiculo c/iva"},
                    {"46": "Gastos de vehiculo S/iva"},
                    {"47": "combustible"},
                    {"48": "fletes c/iva"},
                    {"49": "fletes s/iva"},
                    {"50": "alquiler con iva"},
                    {"51": "alquiler sin iva"},
                    {"52": "gsts ch/rechazados"},
                    {"53": "comisiones pagadas"},
                    {"54": "mant y rep bs. De uso"},
                    {"55": "mant y rep edificio"},
                    {"56": "alquiler maquinarias"},
                    {"57": "descuentos otorgados"},
                    {"58": "indumentaria"},
                    {"59": "anticipo de materiales"},
                    {"60": "hipoteca"},
                    {"61": "comitentes"},
                    {"62": "inmobiliario"},
                    {"63": "descuentos obtenidos"},
                    {"64": "rodados"},
                    {"65": "instalaciones"},
                    {"66": "maquinarias"},
                    {"67": "sistemas informaticos"},
                    {"68": "compra de mue y utiles c/iva"},
                    {"69": "compra de mue y utiles s/iva"},
                    {"70": "compra de bs uso c/iva"},
                    {"71": "compra de bs de uso s/iva"},
                    {"72": "mejoras"},
                    {"73": "vacunos"},
                    {"74": "equinos"},
                    {"75": "conejos"},
                    {"76": "gallinas ponedoras"},
                    {"77": "mejoras inmuebles ajenos"},
                    {"78": "moldes y matrices"},
                    {"79": "restitucion de gastos"},
                    {"80": "venta de mercaderia c/iva"},
                    {"81": "venta de mercaderia s/iva"},
                    {"82": "venta cons final"},
                    {"83": "venta resumen del dia"},
                    {"84": "venta bs de uso c/iva"},
                    {"85": "venta bs de uso s/iva"},
                    {"86": "venta de combustible"},
                    {"87": "prestacion de servicios"},
                    {"88": "honorarios c/iva"},
                    {"89": "honorarios s/iva"},
                    {"90": "alquiler inmuebles"},
                    {"91": "alquiler de vehiculos"},
                    {"92": "comisiones cobradas"},
                    {"93": "liquidacion verduleria"},
                    {"94": "liquidacion carniceria"},
                    {"95": "liquidacion panaderia"},
                    {"96": "montajes"},
                    {"97": "venta mayorista"},
                    {"98": "fabricacion"},
                    {"99": "ch/rechazados"},
                    {"100": "intereses por prestamos"},
                    {"101": "recibo anulado"},
                    {"102": "venta de maquinarias"},
                    {"103": "liquidacion perfumeria"},
                    {"104": "venta ganado x cta de 3°"},
                    {"105": "pastaje"},
                    {"106": "venta de exportacion"},
                    {"107": "liquidacion agropecuaria"},
                    {"109": "toros"},
                    {"110": "licencia"},
                    {"111": "donaciones"},
                    {"112": "diferencia de cambio"},
                    {"113": "gastos financieros"},
                    {"114": "arrendamientos"},
                    {"115": "dto. De valores"},
                    {"116": "mano de obra de 3°"},
                    {"117": "negociacion de valores"},
                    {"118": "liquidacion dto de cheques"},
                    {"119": "patentes de vehiculos"},
                    {"120": "transporte"},
                    {"121": "alquiler particular c/iva"},
                    {"122": "alquiler particular s/iva"},
                    {"123": "alquiler comercial c/iva"},
                    {"124": "alquiler comercial s/iva"},
                    {"125": "utiles y herramientas"},
                    {"126": "premios"},
                    {"127": "reconocimientos"},
                    {"128": "publicidad y propaganda"},
                    {"129": "gastos de seguridad"},
                    {"130": "servivio de transporte"},
                    {"131": "comprobante anulado"},
                    {"132": "anticipos"},
                    {"133": "gastos carrera"},
                    {"134": "Mejora inmueble Propio"},
                    {"135": "Gastos de medicina"},
                    {"136": "Tasa de Fondeadero"},
                    {"137": "Seguros Leasing"},
                    {"138": "Repuestos e Insumos"},
                    {"139": "Ofrendas y Limosnas"},
                    {"140": "Gastos de Comedor"},
                    {"141": "Ganado propio c/iva"},
                    {"142": "Compra de Ganado"},
                    {"143": "Vta. Carne Vacuna"},
                    {"144": "Envases y Accesorios"},
                    {"145": "Venta de Ganado"},
                    {"146": "Ajuste Contable"},
                    {"147": "Fondo de Comercio"},
                    {"148": "Servicios Personales"},
                    {"149": "COMPRA DE CARNE"},
                    {"150": "Insumos Papas"},
                    {"151": "Gastos de Arrendamiento"},
                    {"152": "Venta de Vehiculo"},
                    {"153": "Compra de Vehiculo"},
                    {"154": "Obras en Curso"},
                    {"155": "Boletos y Pasajes"},
                    {"156": "Alquiler Barco"},
                    {"157": "Materiales de Decoracion"},
                    {"158": "Alquiler y Expensas"},
                    {"159": "Alquiler de Herramientas"},
                    {"160": "Viandas"},
                    {"161": "Intereses"},
                    {"162": "Seguros de Caucion"},
                    {"163": "Impresiones"},
                    {"164": "Gastos de Producción"},
                    {"165": "Prestadores"},
                    {"166": "Devolucion de Mercaderias"},
                    {"167": "Alquiler de Maquinarias"},
                    {"168": "Certificados Revisión Técnica"},
                    {"169": "Registro Control Modelo"},
                    {"170": "Camara Arg. De Talleres"},
                    {"171": "Honorarios Directores"},
                    {"172": "Fondo de Reparo"},
                    {"173": "Gastos de Sanidad"},
                    {"174": "Plan de ahorro"},
                    {"175": "Alquiler Temporario"},
                    {"176": "Alquiler y Logistica"},
                    {"177": "Alquiler Bs. Muebles"},
                    {"178": "Gastos de Capacitación"},
                    {"179": "Maq. Y equipos medicos"},
                    {"180": "gastos de organización"},
                    {"181": "equipos de comunicación"},
                    {"182": "Gas para la venta"},
                    {"183": "Venta Flete Internacional"},
                    {"184": "Flete Internacional"},
                    {"185": "Gastos de Obra"},
                    {"186": "Gastos de Desarrollo"},
                    {"187": "Embarcaciones"},
                    {"188": "Gastos de embarcacion"},
                    {"189": "Venta de Papa"},
                    {"190": "Utiles y elementos de cocina"},
                    {"191": "cubiertos y vajillas"},
                    {"192": "elementos ortopedicos"},
                    {"193": "pines"},
                    {"194": "golosinas"},
                    {"195": "rotary internacional"},
                    {"196": "distrito rotario 4825"},
                    {"197": "ret- seguridad e higiene a"},
                    {"198": "gastos de representacion"},
                    {"199": "ativo de caja  (compra + v"},
                    {"201": "C.M."},
                    {"202": "ativo de caja  (compra + v"},
                    {"203": "Alimentos"},
                    {"204": "Enfriado"},
                ]

                try:
                    encabezado = lines[1:7]
                    primer_fecha = ""
                    # Eliminar los saltos de línea y caracteres no deseados
                    encabezado_limpio = [
                        line.replace("\n", "").strip() for line in encabezado
                    ]

                    # Intentar crear el diccionario
                    encabezado_completo = {
                        "RAZON SOCIAL": encabezado_limpio[0],
                        "DIRECCION": encabezado_limpio[1],
                        "CUIT": encabezado_limpio[2],
                        "LIBRO": encabezado_limpio[3].split("  ")[-1],
                        "PERIODO": encabezado_limpio[4].split("  ")[-1],
                    }

                except Exception as e:
                    print(f"Ocurrió un error al procesar el encabezado: {e}")
                    encabezado_completo = {}  # No se crea el objeto si hay un error

                for i, line in enumerate(lines[9:], start=2):
                    # Si la línea contiene "TOTALES POR TASA", eliminar todo lo posterior
                    if "IVA VENTAS" in line:
                        compras_o_ventas = "Ventas"
                    elif "IVA COMPRAS" in line:
                        compras_o_ventas = "Compras"

                    if "TOTALES POR TASA" in line:
                        eliminar_desde_totales = True
                        continue  # Saltar esta línea (no la procesamos, solo marcamos la eliminación)

                    if eliminar_desde_totales:
                        break  # Detener el procesamiento de líneas después de "TOTALES POR TASA"

                    # Si la línea comienza con "----", marca el inicio del bloque a eliminar
                    if line.startswith("----"):
                        eliminar = True
                        continue  # Saltar esta línea (ya no se procesa)

                    # Si la línea comienza con "--", marca el final del bloque a eliminar
                    if line.startswith("--"):
                        eliminar = False
                        continue  # Saltar esta línea (ya no se procesa)

                    # Si no estamos dentro de un bloque a eliminar, limpiamos la línea
                    if not eliminar:
                        cleaned_line = re.sub(
                            r"\x1b[^m]*m", "", line
                        )  # Elimina secuencias ANSI
                        cleaned_line = re.sub(
                            r"[\x00-\x1F\x7F]", "", cleaned_line
                        )  # Eliminación de caracteres de control ASCII
                        if "Ñ" in cleaned_line:
                            if len(cleaned_line) == 135:
                                pass
                            else:
                                cleaned_line = (
                                    cleaned_line[:44] + " " + cleaned_line[44:]
                                )
                        cleaned_lines.append(
                            cleaned_line
                        )  # Guardar la línea limpia sin espacios extra

                doble_cleaned_lines = []

                for i, line in enumerate(cleaned_lines, start=-1):
                    if "PPag." in line:
                        break
                    doble_cleaned_lines.append(line)

                for index, cleaned_line in enumerate(doble_cleaned_lines):
                    if (
                        "numero" in temp_movement
                        and temp_movement["numero"] == cleaned_line[12:20]
                    ):
                        partes = re.split(r"\s{3,}", cleaned_line[70:])
                        if len(partes) < 2:
                            pass
                        else:
                            if partes[0] == "Tasa 21%":
                                temp_movement[partes[0] + " Neto"] = partes[1]
                                temp_movement[partes[0] + " IVA"] = partes[2]
                            elif partes[0] == "T.10.5%":
                                temp_movement[partes[0] + " Neto"] = partes[1]
                                temp_movement[partes[0] + " IVA"] = partes[2]
                            elif partes[0] == "Tasa 27%":
                                temp_movement[partes[0] + " Neto"] = partes[1]
                                temp_movement[partes[0] + " IVA"] = partes[2]
                            elif partes[0] == "C.F.21%":
                                temp_movement[partes[0] + " Neto"] = partes[1]
                                temp_movement[partes[0] + " IVA"] = partes[2]
                            elif partes[0] == "C.F.10.5%":
                                temp_movement[partes[0] + " Neto"] = partes[1]
                                temp_movement[partes[0] + " IVA"] = partes[2]
                            elif partes[0] == "Tasa 2.5%":
                                temp_movement[partes[0] + " Neto"] = partes[1]
                                temp_movement[partes[0] + " IVA"] = partes[2]
                            elif partes[0] == "T.IMP 21%":
                                temp_movement[partes[0] + " Neto"] = partes[1]
                                temp_movement[partes[0] + " IVA"] = partes[2]
                            elif partes[0] == "T.IMP 10%":
                                temp_movement[partes[0] + " Neto"] = partes[1]
                                temp_movement[partes[0] + " IVA"] = partes[2]
                            elif partes[0] == "R.Monot21":
                                if compras_o_ventas == "Ventas":
                                    temp_movement[partes[0] + " Neto"] = partes[1]
                                    temp_movement[partes[0] + " IVA"] = partes[2]
                                else:
                                    temp_movement[partes[0]] = partes[1]
                            elif partes[0] == "R.Mont.10":
                                if compras_o_ventas == "Ventas":
                                    temp_movement[partes[0] + " Neto"] = partes[1]
                                    temp_movement[partes[0] + " IVA"] = partes[2]
                                else:
                                    temp_movement[partes[0]] = partes[1]
                    else:
                        if cleaned_line[0:2] == "  ":
                            partes = re.split(r"\s{3,}", cleaned_line[70:])
                            if len(partes) < 2:
                                pass
                            else:
                                if partes[0] == "Tasa 21%":
                                    if partes[0] + " Neto" in temp_movement:
                                        neto_numero_anterior = float(
                                            temp_movement[partes[0] + " Neto"].replace(
                                                ",", "."
                                            )
                                        )
                                        iva_numero_anterior = float(
                                            temp_movement[partes[0] + " IVA"].replace(
                                                ",", "."
                                            )
                                        )
                                        neto = float(partes[1].replace(",", "."))
                                        iva = float(partes[2].replace(",", "."))
                                        temp_movement[partes[0] + " Neto"] = round(
                                            neto_numero_anterior + neto, 2
                                        )
                                        temp_movement[partes[0] + " IVA"] = round(
                                            iva_numero_anterior + iva, 2
                                        )
                                    else:
                                        temp_movement[partes[0] + " Neto"] = partes[1]
                                        temp_movement[partes[0] + " IVA"] = partes[2]
                                elif partes[0] == "T.10.5%":
                                    if partes[0] + " Neto" in temp_movement:
                                        neto_numero_anterior = float(
                                            temp_movement[partes[0] + " Neto"].replace(
                                                ",", "."
                                            )
                                        )
                                        iva_numero_anterior = float(
                                            temp_movement[partes[0] + " IVA"].replace(
                                                ",", "."
                                            )
                                        )
                                        neto = float(partes[1].replace(",", "."))
                                        iva = float(partes[2].replace(",", "."))
                                        temp_movement[partes[0] + " Neto"] = round(
                                            neto_numero_anterior + neto, 2
                                        )
                                        temp_movement[partes[0] + " IVA"] = round(
                                            iva_numero_anterior + iva, 2
                                        )
                                    else:
                                        temp_movement[partes[0] + " Neto"] = partes[1]
                                        temp_movement[partes[0] + " IVA"] = partes[2]
                                elif partes[0] == "Tasa 27%":
                                    if partes[0] + " Neto" in temp_movement:
                                        neto_numero_anterior = float(
                                            temp_movement[partes[0] + " Neto"].replace(
                                                ",", "."
                                            )
                                        )
                                        iva_numero_anterior = float(
                                            temp_movement[partes[0] + " IVA"].replace(
                                                ",", "."
                                            )
                                        )
                                        neto = float(partes[1].replace(",", "."))
                                        iva = float(partes[2].replace(",", "."))
                                        temp_movement[partes[0] + " Neto"] = round(
                                            neto_numero_anterior + neto, 2
                                        )
                                        temp_movement[partes[0] + " IVA"] = round(
                                            iva_numero_anterior + iva, 2
                                        )
                                    else:
                                        temp_movement[partes[0] + " Neto"] = partes[1]
                                        temp_movement[partes[0] + " IVA"] = partes[2]
                                elif partes[0] == "C.F.21%":
                                    if partes[0] + " Neto" in temp_movement:
                                        neto_numero_anterior = float(
                                            temp_movement[partes[0] + " Neto"].replace(
                                                ",", "."
                                            )
                                        )
                                        iva_numero_anterior = float(
                                            temp_movement[partes[0] + " IVA"].replace(
                                                ",", "."
                                            )
                                        )
                                        neto = float(partes[1].replace(",", "."))
                                        iva = float(partes[2].replace(",", "."))
                                        temp_movement[partes[0] + " Neto"] = round(
                                            neto_numero_anterior + neto, 2
                                        )
                                        temp_movement[partes[0] + " IVA"] = round(
                                            iva_numero_anterior + iva, 2
                                        )
                                    else:
                                        temp_movement[partes[0] + " Neto"] = partes[1]
                                        temp_movement[partes[0] + " IVA"] = partes[2]
                                elif partes[0] == "C.F.10.5%":
                                    if partes[0] + " Neto" in temp_movement:
                                        neto_numero_anterior = float(
                                            temp_movement[partes[0] + " Neto"].replace(
                                                ",", "."
                                            )
                                        )
                                        iva_numero_anterior = float(
                                            temp_movement[partes[0] + " IVA"].replace(
                                                ",", "."
                                            )
                                        )
                                        neto = float(partes[1].replace(",", "."))
                                        iva = float(partes[2].replace(",", "."))
                                        temp_movement[partes[0] + " Neto"] = round(
                                            neto_numero_anterior + neto, 2
                                        )
                                        temp_movement[partes[0] + " IVA"] = round(
                                            iva_numero_anterior + iva, 2
                                        )
                                    else:
                                        temp_movement[partes[0] + " Neto"] = partes[1]
                                        temp_movement[partes[0] + " IVA"] = partes[2]
                                elif partes[0] == "Tasa 2.5%":
                                    if partes[0] + " Neto" in temp_movement:
                                        neto_numero_anterior = float(
                                            temp_movement[partes[0] + " Neto"].replace(
                                                ",", "."
                                            )
                                        )
                                        iva_numero_anterior = float(
                                            temp_movement[partes[0] + " IVA"].replace(
                                                ",", "."
                                            )
                                        )
                                        neto = float(partes[1].replace(",", "."))
                                        iva = float(partes[2].replace(",", "."))
                                        temp_movement[partes[0] + " Neto"] = round(
                                            neto_numero_anterior + neto, 2
                                        )
                                        temp_movement[partes[0] + " IVA"] = round(
                                            iva_numero_anterior + iva, 2
                                        )
                                    else:
                                        temp_movement[partes[0] + " Neto"] = partes[1]
                                        temp_movement[partes[0] + " IVA"] = partes[2]
                                elif partes[0] == "T.IMP 21%":
                                    if partes[0] + " Neto" in temp_movement:
                                        neto_numero_anterior = float(
                                            temp_movement[partes[0] + " Neto"].replace(
                                                ",", "."
                                            )
                                        )
                                        iva_numero_anterior = float(
                                            temp_movement[partes[0] + " IVA"].replace(
                                                ",", "."
                                            )
                                        )
                                        neto = float(partes[1].replace(",", "."))
                                        iva = float(partes[2].replace(",", "."))
                                        temp_movement[partes[0] + " Neto"] = round(
                                            neto_numero_anterior + neto, 2
                                        )
                                        temp_movement[partes[0] + " IVA"] = round(
                                            iva_numero_anterior + iva, 2
                                        )
                                    else:
                                        temp_movement[partes[0] + " Neto"] = partes[1]
                                        temp_movement[partes[0] + " IVA"] = partes[2]
                                elif partes[0] == "T.IMP 10%":
                                    if partes[0] + " Neto" in temp_movement:
                                        neto_numero_anterior = float(
                                            temp_movement[partes[0] + " Neto"].replace(
                                                ",", "."
                                            )
                                        )
                                        iva_numero_anterior = float(
                                            temp_movement[partes[0] + " IVA"].replace(
                                                ",", "."
                                            )
                                        )
                                        neto = float(partes[1].replace(",", "."))
                                        iva = float(partes[2].replace(",", "."))
                                        temp_movement[partes[0] + " Neto"] = round(
                                            neto_numero_anterior + neto, 2
                                        )
                                        temp_movement[partes[0] + " IVA"] = round(
                                            iva_numero_anterior + iva, 2
                                        )
                                    else:
                                        temp_movement[partes[0] + " Neto"] = partes[1]
                                        temp_movement[partes[0] + " IVA"] = partes[2]
                                elif partes[0] == "R.Monot21":
                                    if compras_o_ventas == "Ventas":
                                        if partes[0] + " Neto" in temp_movement:
                                            neto_numero_anterior = float(
                                                temp_movement[
                                                    partes[0] + " Neto"
                                                ].replace(",", ".")
                                            )
                                            iva_numero_anterior = float(
                                                temp_movement[
                                                    partes[0] + " IVA"
                                                ].replace(",", ".")
                                            )
                                            neto = float(partes[1].replace(",", "."))
                                            iva = float(partes[2].replace(",", "."))
                                            temp_movement[partes[0] + " Neto"] = round(
                                                neto_numero_anterior + neto, 2
                                            )
                                            temp_movement[partes[0] + " IVA"] = round(
                                                iva_numero_anterior + iva, 2
                                            )
                                        else:
                                            temp_movement[partes[0] + " Neto"] = partes[
                                                1
                                            ]
                                            temp_movement[partes[0] + " IVA"] = partes[
                                                2
                                            ]
                                    else:
                                        temp_movement[partes[0]] = partes[1]
                                elif partes[0] == "R.Mont.10":
                                    if compras_o_ventas == "Ventas":
                                        if partes[0] + " Neto" in temp_movement:
                                            neto_numero_anterior = float(
                                                temp_movement[
                                                    partes[0] + " Neto"
                                                ].replace(",", ".")
                                            )
                                            iva_numero_anterior = float(
                                                temp_movement[
                                                    partes[0] + " IVA"
                                                ].replace(",", ".")
                                            )
                                            neto = float(partes[1].replace(",", "."))
                                            iva = float(partes[2].replace(",", "."))
                                            temp_movement[partes[0] + " Neto"] = round(
                                                neto_numero_anterior + neto, 2
                                            )
                                            temp_movement[partes[0] + " IVA"] = round(
                                                iva_numero_anterior + iva, 2
                                            )
                                        else:
                                            temp_movement[partes[0] + " Neto"] = partes[
                                                1
                                            ]
                                            temp_movement[partes[0] + " IVA"] = partes[
                                                2
                                            ]
                                    else:
                                        temp_movement[partes[0]] = partes[1]
                                else:
                                    if partes[0] in temp_movement:
                                        if type(temp_movement[partes[0]]) == float:
                                            numero_actual = float(
                                                partes[1].replace(",", ".")
                                            )
                                            temp_movement[partes[0]] = round(
                                                temp_movement[partes[0]]
                                                + numero_actual,
                                                2,
                                            )
                                        else:
                                            numero_anterior = float(
                                                temp_movement[partes[0]].replace(
                                                    ",", "."
                                                )
                                            )
                                            numero_actual = float(
                                                partes[1].replace(",", ".")
                                            )
                                            temp_movement[partes[0]] = round(
                                                numero_anterior + numero_actual, 2
                                            )
                                    else:
                                        temp_movement[partes[0]] = partes[1]
                            if (
                                index == len(doble_cleaned_lines) - 1
                            ):  # Verificar si es el último elemento
                                movements.append(temp_movement)
                        else:
                            partes = re.split(r"\s{3,}", cleaned_line[70:])
                            movement = temp_movement.copy()
                            movements.append(movement)
                            temp_movement.clear()
                            if len(partes) < 2:
                                pass
                            else:
                                concepto = cleaned_line[64:67].strip()
                                descripcion = ""
                                for diccionario in conceptos:
                                    if concepto in diccionario:
                                        descripcion = diccionario[concepto]
                                        break
                                jurisdicciones.add(cleaned_line[68:69])
                                temp_movement = {
                                    "Fecha": cleaned_line[0:2],
                                    "Comprobante": cleaned_line[3:5],
                                    "PV": cleaned_line[6:11],
                                    "Nro": cleaned_line[12:20],
                                    "Letra": cleaned_line[20:21],
                                    "Razon Social": cleaned_line[22:44],
                                    "Condicion": cleaned_line[45:49],
                                    "CUIT": cleaned_line[50:63],
                                    "Concepto": cleaned_line[64:67],
                                    "Descripcion": descripcion.upper(),
                                    "Jurisdiccion": cleaned_line[68:69],
                                }
                                if len(partes) == 3:
                                    primer_monto = partes[1].split(" ")
                                    primer_monto = list(filter(None, primer_monto))
                                    segundo_monto = partes[1].split(" ")
                                    segundo_monto = list(filter(None, segundo_monto))
                                    partes = [partes[0]] + primer_monto + segundo_monto

                                if partes[0] == "Tasa 21%":
                                    temp_movement[partes[0] + " Neto"] = partes[1]
                                    temp_movement[partes[0] + " IVA"] = partes[2]
                                elif partes[0] == "T.10.5%":
                                    temp_movement[partes[0] + " Neto"] = partes[1]
                                    temp_movement[partes[0] + " IVA"] = partes[2]
                                elif partes[0] == "Tasa 27%":
                                    temp_movement[partes[0] + " Neto"] = partes[1]
                                    temp_movement[partes[0] + " IVA"] = partes[2]
                                elif partes[0] == "C.F.21%":
                                    temp_movement[partes[0] + " Neto"] = partes[1]
                                    temp_movement[partes[0] + " IVA"] = partes[2]
                                elif partes[0] == "C.F.10.5%":
                                    temp_movement[partes[0] + " Neto"] = partes[1]
                                    temp_movement[partes[0] + " IVA"] = partes[2]
                                elif partes[0] == "Tasa 2.5%":
                                    temp_movement[partes[0] + " Neto"] = partes[1]
                                    temp_movement[partes[0] + " IVA"] = partes[2]
                                elif partes[0] == "T.IMP 21%":
                                    temp_movement[partes[0] + " Neto"] = partes[1]
                                    temp_movement[partes[0] + " IVA"] = partes[2]
                                elif partes[0] == "T.IMP 10%":
                                    temp_movement[partes[0] + " Neto"] = partes[1]
                                    temp_movement[partes[0] + " IVA"] = partes[2]
                                elif partes[0] == "R.Monot21":
                                    if compras_o_ventas == "Ventas":
                                        temp_movement[partes[0] + " Neto"] = partes[1]
                                        temp_movement[partes[0] + " IVA"] = partes[2]
                                    else:
                                        temp_movement[partes[0]] = partes[1]
                                elif partes[0] == "R.Mont.10":
                                    if compras_o_ventas == "Ventas":
                                        temp_movement[partes[0] + " Neto"] = partes[1]
                                        temp_movement[partes[0] + " IVA"] = partes[2]
                                    else:
                                        temp_movement[partes[0]] = partes[1]
                                else:
                                    temp_movement[partes[0]] = partes[1]
                                if (
                                    index == len(doble_cleaned_lines) - 1
                                ):  # Verificar si es el último elemento
                                    movements.append(temp_movement)

                if not movements[0]:
                    movements.pop(0)

                df_encabezado = pd.DataFrame(
                    list(encabezado_completo.values()), columns=["Valor"]
                )

                # Crear el DataFrame
                df = pd.DataFrame(movements)
                # Reemplazar los valores NaN por 0
                df = df.fillna(0)

                # Reemplazar comas por puntos en las columnas numéricas (desde el índice 11 hasta el final)
                df.iloc[:, 11:] = df.iloc[:, 11:].replace(",", ".", regex=True)

                # Convertir las columnas desde el índice 11 hasta el final a tipo numérico
                df.iloc[:, 11:] = (
                    df.iloc[:, 11:].apply(pd.to_numeric, errors="coerce").fillna(0)
                )

                # Lista de columnas que deseas volver negativas
                columnas_a_convertir = df.columns[11:]

                # Aplicar la conversión a negativos si el tipo de comprobante es "NC"
                df.loc[df["Comprobante"] == "NC", columnas_a_convertir] *= -1

                df["PV"] = pd.to_numeric(df["PV"])
                df["Nro"] = pd.to_numeric(df["Nro"])
                df["Concepto"] = pd.to_numeric(df["Concepto"])

                # Crear una lista para almacenar las filas combinadas
                resultado = []

                # Variable para almacenar la fila combinada actual
                fila_actual = df.iloc[0].copy()

                for i in range(1, len(df)):
                    fila_siguiente = df.iloc[i]

                    # Si la clave principal se repite, combinar valores
                    if (
                        fila_actual["Nro"] == fila_siguiente["Nro"]
                        and fila_actual["PV"] == fila_siguiente["PV"]
                        and fila_actual["Razon Social"]
                        == fila_siguiente["Razon Social"]
                    ):
                        for col in df.columns[11:]:  # Sumar solo las columnas numéricas
                            fila_actual[col] += fila_siguiente[col]
                    else:
                        resultado.append(fila_actual)
                        fila_actual = fila_siguiente.copy()

                # Agregar la última fila combinada
                resultado.append(fila_actual)

                # Convertir lista a DataFrame
                df_final = pd.DataFrame(resultado)

                df_final["Total"] = df_final.iloc[:, 11:].sum(axis=1)

                # Crear una fila con la etiqueta "TOTAL"
                fila_total = pd.DataFrame(df_final.iloc[:, 11:].sum()).T

                # Agregar un identificador en la primera columna para que se distinga la fila de totalización
                fila_total.insert(0, "Nro", "TOTAL")
                fila_total.insert(1, "Razon Social", "")

                # Concatenar la fila total con el DataFrame
                df_final = pd.concat([df_final, fila_total], ignore_index=True)

                # Guardar el DataFrame en un archivo Excel
                excel_filename = "Movimientos.xlsx"

                # Agrupar por 'Concepto' y sumar los valores numéricos
                df_grouped = df_final.groupby("Concepto", as_index=False).sum(
                    numeric_only=True
                )

                # Filtrar columnas, eliminando las que contienen ciertas palabras
                df_grouped = df_grouped.loc[
                    :,
                    ~df_grouped.columns.str.contains(
                        "IVA|RET|PERC|Total|PV|SIRCREB|RG", case=False
                    ),
                ]

                # Excluir la primera columna (posición 0) de la suma
                columnas_numericas = df_grouped.select_dtypes(include="number").columns
                columnas_a_sumar = columnas_numericas[
                    1:
                ]  # Todas las numéricas excepto la primera

                # Sumar solo las columnas seleccionadas
                df_grouped["Neto"] = df_grouped[columnas_a_sumar].sum(axis=1)

                # Crear una fila con la etiqueta "TOTAL" sumando todas las columnas numéricas
                fila_total = pd.DataFrame(
                    df_grouped.iloc[:, 1:].sum()
                ).T  # Excluye la primera columna al sumar

                # Asegurar que la fila total tenga una etiqueta en "Concepto"
                fila_total.insert(
                    0, df_grouped.columns[0], "TOTAL"
                )  # Agregar "TOTAL" a la primera columna

                # Concatenar la fila total con el DataFrame
                df_grouped = pd.concat([df_grouped], ignore_index=True)

                # Convertir ambas columnas 'Concepto' a tipo float64
                df_grouped["Concepto"] = df_grouped["Concepto"].astype(float)
                df_final["Concepto"] = df_final["Concepto"].astype(float)

                # Hacer el merge entre df1 y df2
                df_merged = pd.merge(
                    df_grouped,  # Tabla 2
                    df_final[
                        ["Concepto", "Descripcion"]
                    ],  # Tabla 1: solo las columnas Concepto y Descripcion
                    on="Concepto",  # Columna clave para hacer el join
                    how="left",  # El tipo de merge será 'left' para conservar todos los registros de la tabla 2
                )

                df_merged = df_merged.drop_duplicates(subset="Concepto", keep="first")

                df_merged = df_merged[["Concepto", "Descripcion", "Neto"]]

                df_final = df_final.drop("Descripcion", axis=1)

                # Eliminar columnas que contienen "Unnamed"
                df_final = df_final.loc[:, ~(df_final.columns == "")]

                # Filtrar columnas, eliminando las que contienen ciertas palabras
                df_netos = df_final.loc[
                    :,
                    ~df_final.columns.str.contains(
                        "IVA|RET|PERC|Total|SIRCREB|RG", case=False
                    ),
                ]

                # Excluir la primera columna (posición 0) de la suma
                columnas_numericas_a_sumar = df_netos.select_dtypes(
                    include="number"
                ).columns
                columnas_a_sumar = columnas_numericas_a_sumar[10:]

                df_netos["Total NETO"] = df_netos.iloc[:, 10:].sum(axis=1)

                # Crear una fila con la etiqueta "TOTAL"
                fila_total = pd.DataFrame(df_netos.iloc[:, 10:].sum()).T

                df_encabezado.columns = [""] * len(df_encabezado.columns)

                # Crear una copia
                df_total_por_jurisdiccion_y_concepto = df_merged.drop(
                    "Neto", axis=1
                ).copy()

                # Obtener los rangos de las columnas a formular
                inicio = 10
                fin = inicio + len(df_netos) - 2

                ultima_columna_index = df_netos.columns.get_loc(
                    "Total NETO"
                )  # Índice numérico de la columna
                ultima_columna_letra = chr(
                    65 + ultima_columna_index
                )  # Convertir a letra (solo para A-Z)

                rango_concepto = f"$I${inicio}:$I${fin}"
                rango_jurisdiccion = f"$J${inicio}:$J${fin}"
                rango_total_neto = (
                    f"${ultima_columna_letra}${inicio}:${ultima_columna_letra}${fin}"
                )

                for jurisdiccion in jurisdicciones:
                    df_total_por_jurisdiccion_y_concepto[jurisdiccion] = (
                        0  # Puedes cambiar este valor por otro si es necesario
                    )

                jurisdicciones = sorted(jurisdicciones)

                # Crear un relleno sólido (color verde en este caso)
                relleno_verde = PatternFill(
                    start_color="5CE65C", end_color="5CE65C", fill_type="solid"
                )

                with pd.ExcelWriter(excel_filename, engine="openpyxl") as writer:
                    df_encabezado.to_excel(
                        writer, sheet_name="Netos", startcol=5, index=False
                    )
                    df_netos.to_excel(
                        writer, sheet_name="Netos", startrow=8, index=False
                    )

                    df_total_por_jurisdiccion_y_concepto.to_excel(
                        writer,
                        sheet_name="Netos",
                        startrow=fin + 4,
                        startcol=6,
                        index=False,
                    )

                    df_encabezado.to_excel(
                        writer, sheet_name="Movimientos", startcol=5, index=False
                    )
                    df_final.to_excel(
                        writer, sheet_name="Movimientos", startrow=8, index=False
                    )

                    df_merged.to_excel(
                        writer,
                        sheet_name="Netos",
                        startrow=fin + 4,
                        startcol=2,
                        index=False,
                    )

                    df_merged.to_excel(
                        writer,
                        sheet_name="CONCEPTOS",
                        index=False,
                    )

                # 🔹 **Abrir el archivo con `openpyxl` para insertar las fórmulas**

                wb = load_workbook(excel_filename)
                ws = wb["Netos"]

                # 🔹 **Insertar las fórmulas directamente en la hoja de cálculo**

                contador = 0
                contador_concepto = 0

                # CREACION DE CELDAS CON FORMULA EN TABLA TOTALES POR CONCEPTOS

                for index, row in df_merged.iterrows():
                    fila_excel_concepto = fin + 6
                    fila_temporal = contador_concepto + fila_excel_concepto

                    formula = (
                        f"=SUMIF({rango_concepto},$G{fila_temporal},{rango_total_neto})"
                    )
                    col_letter = get_column_letter(5)
                    ws[f"{col_letter}{fila_temporal}"] = formula

                    contador_concepto += 1

                # CREACION DE CELDAS CON FORMULA EN TABLA 3

                for index, row in df_total_por_jurisdiccion_y_concepto.iterrows():
                    fila_excel_concepto = fin + 6
                    fila_temporal = contador + fila_excel_concepto
                    fila_excel_jurisdiccion = fin + 5
                    for col_index, jurisdiccion in enumerate(
                        jurisdicciones, start=9
                    ):  # "D" es la 4ta columna
                        col_letter = get_column_letter(
                            col_index
                        )  # Convertir índice numérico a letra
                        formula = f"=SUMIFS({rango_total_neto},{rango_concepto},$G{fila_temporal},{rango_jurisdiccion},{col_letter}${fila_excel_jurisdiccion})"
                        ws[f"{col_letter}{fila_temporal}"] = formula  # Asignar fórmula
                    contador += 1

                ultima_fila = fin + 1
                total_columnas = sum(1 for celda in ws[9] if celda.value is not None)
                col_letter_color = get_column_letter(total_columnas)
                # Aplicar el color a la última celda
                celda_a_pintar = ws[f"{col_letter_color}{ultima_fila}"]
                celda_a_pintar.fill = relleno_verde
                celda_a_pintar.font = Font(bold=True, size=12)

                col_letter = get_column_letter(8)
                ws[f"{col_letter}{fin + 4}"] = "TOTALES POR CONCEPTO Y JURISDICCION"
                ws[f"{col_letter}{fin + 4}"].font = Font(bold=True, size=14)

                col_letter_concepto = get_column_letter(4)
                ws[f"{col_letter_concepto}{fin + 4}"] = "TOTALES POR CONCEPTO"
                ws[f"{col_letter_concepto}{fin + 4}"].font = Font(bold=True, size=14)

                wb.save(excel_filename)

                # Enviar el archivo Excel generado
                return send_from_directory(
                    os.getcwd(), excel_filename, as_attachment=True
                )

            except Exception as e:
                return f"Hubo un error al procesar el archivo: {e}"

    return render_template("index.html")


if __name__ == "__main__":
    if not os.path.exists(app.config["UPLOAD_FOLDER"]):
        os.makedirs(app.config["UPLOAD_FOLDER"])
    app.run(debug=True, host="0.0.0.0", port=os.getenv("PORT", default=5000))
